"""
Premium Pharma-Grade Report Export — Excel (.xlsx) and PDF (.pdf).

Design: Clean, professional pharma styling.
Performance: Pre-created style objects, range-based operations, single-pass rendering.
"""


def generate_excel_report(columns, table_data, file_path, header_info, currency_columns=None, totals_row=None):
    """
    Generate a premium Excel report using openpyxl with pre-created styles.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = header_info.get("report_title", "Report")[:31]  # Sheet name max 31 chars

        num_cols = len(columns)
        end_col_letter = get_column_letter(num_cols)

        # ── Pre-create all style objects (reusable, no per-cell creation) ──
        PRIMARY = "2F75B5"
        SECONDARY = "D9E1F2"
        WHITE = "FFFFFF"
        DARK = "1F2937"
        LIGHT_GRAY = "F2F2F2"

        company_font = Font(name="Calibri", size=14, bold=True, color=WHITE)
        sub_header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        title_font = Font(name="Calibri", size=12, bold=True, color=PRIMARY)
        date_font = Font(name="Calibri", size=10, italic=True, color="555555")
        filter_font = Font(name="Calibri", size=9, color="666666")
        table_header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        data_font = Font(name="Calibri", size=10, color=DARK)
        totals_font = Font(name="Calibri", size=10, bold=True, color=PRIMARY)
        currency_font = Font(name="Calibri", size=10, color=DARK)

        primary_fill = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
        secondary_fill = PatternFill(start_color=SECONDARY, end_color=SECONDARY, fill_type="solid")
        alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
        white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
        totals_fill = PatternFill(start_color=SECONDARY, end_color=SECONDARY, fill_type="solid")

        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        thin_side = Side(style="thin", color="BFBFBF")
        header_border = Border(bottom=Side(style="medium", color=PRIMARY))
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        totals_border = Border(top=Side(style="double", color=PRIMARY), bottom=Side(style="double", color=PRIMARY),
                               left=thin_side, right=thin_side)

        # Identify currency column indices
        currency_col_indices = set()
        if currency_columns:
            for i, col in enumerate(columns):
                if col in currency_columns:
                    currency_col_indices.add(i)

        # ── HEADER SECTION ──
        # Row 1: Company Name (dark blue background)
        ws.merge_cells(f"A1:{end_col_letter}1")
        c = ws["A1"]
        c.value = header_info.get("company_name", "PharmIQ")
        c.font = company_font
        c.fill = primary_fill
        c.alignment = center_align
        ws.row_dimensions[1].height = 30

        # Row 2: Address
        ws.merge_cells(f"A2:{end_col_letter}2")
        c = ws["A2"]
        c.value = header_info.get("address", "")
        c.font = sub_header_font
        c.fill = primary_fill
        c.alignment = center_align
        ws.row_dimensions[2].height = 18

        # Row 3: Contact + GST
        contact_parts = []
        if header_info.get("phone"):
            contact_parts.append(f"Phone: {header_info['phone']}")
        if header_info.get("email"):
            contact_parts.append(f"Email: {header_info['email']}")
        if header_info.get("gst_no"):
            contact_parts.append(f"GST: {header_info['gst_no']}")
        ws.merge_cells(f"A3:{end_col_letter}3")
        c = ws["A3"]
        c.value = " | ".join(contact_parts) if contact_parts else ""
        c.font = sub_header_font
        c.fill = primary_fill
        c.alignment = center_align
        ws.row_dimensions[3].height = 18

        # Row 4: Licence No
        ws.merge_cells(f"A4:{end_col_letter}4")
        c = ws["A4"]
        c.value = f"Licence No.: {header_info.get('licence_no', '')}" if header_info.get("licence_no") else ""
        c.font = sub_header_font
        c.fill = primary_fill
        c.alignment = center_align
        ws.row_dimensions[4].height = 18

        # Row 5: Spacer
        ws.merge_cells(f"A5:{end_col_letter}5")
        ws.row_dimensions[5].height = 6

        # Row 6: Report Title
        ws.merge_cells(f"A6:{end_col_letter}6")
        c = ws["A6"]
        c.value = header_info.get("report_title", "Report")
        c.font = title_font
        c.alignment = center_align
        ws.row_dimensions[6].height = 25

        # Row 7: Date Range
        ws.merge_cells(f"A7:{end_col_letter}7")
        c = ws["A7"]
        c.value = header_info.get("date_range", "")
        c.font = date_font
        c.alignment = center_align
        ws.row_dimensions[7].height = 18

        # Row 8: Filter info (if customer or medicine specified)
        filter_parts = []
        if header_info.get("customer_filter"):
            filter_parts.append(f"Customer: {header_info['customer_filter']}")
        if header_info.get("medicine_filter"):
            filter_parts.append(f"Medicine: {header_info['medicine_filter']}")
        ws.merge_cells(f"A8:{end_col_letter}8")
        c = ws["A8"]
        c.value = " | ".join(filter_parts) if filter_parts else ""
        c.font = filter_font
        c.alignment = center_align
        ws.row_dimensions[8].height = 16

        # Row 9: Spacer
        ws.merge_cells(f"A9:{end_col_letter}9")
        ws.row_dimensions[9].height = 6

        # ── TABLE HEADER (Row 10) ──
        header_row_idx = 10
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=header_row_idx, column=col_idx, value=col_name)
            cell.font = table_header_font
            cell.fill = primary_fill
            cell.alignment = center_align
            cell.border = header_border
        ws.row_dimensions[header_row_idx].height = 28

        # ── TABLE DATA ──
        current_row = header_row_idx + 1
        col_max_widths = [len(str(col)) for col in columns]  # Track for auto-width

        for i, row_dict in enumerate(table_data):
            values = row_dict.get("values", [])
            is_even = (i % 2 == 0)
            row_fill = alt_fill if is_even else white_fill

            for col_idx in range(num_cols):
                val = values[col_idx] if col_idx < len(values) else ""
                cell = ws.cell(row=current_row, column=col_idx + 1, value=val)
                cell.font = data_font
                cell.fill = row_fill
                cell.border = cell_border
                cell.alignment = right_align if col_idx in currency_col_indices else left_align

                # Track max width
                val_len = len(str(val)) if val else 0
                if val_len > col_max_widths[col_idx]:
                    col_max_widths[col_idx] = val_len

            current_row += 1

        # ── TOTALS ROW ──
        if totals_row:
            for col_idx in range(num_cols):
                val = totals_row[col_idx] if col_idx < len(totals_row) else ""
                cell = ws.cell(row=current_row, column=col_idx + 1, value=val)
                cell.font = totals_font
                cell.fill = totals_fill
                cell.border = totals_border
                cell.alignment = right_align if col_idx in currency_col_indices else left_align
            ws.row_dimensions[current_row].height = 28
            current_row += 1

        # ── AUTO COLUMN WIDTHS (single pass, already computed) ──
        for col_idx in range(num_cols):
            width = min(col_max_widths[col_idx] + 4, 40)
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = max(width, 12)

        # ── PRINT SETTINGS ──
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        wb.save(file_path)
        return True, "Excel report generated successfully."
    except Exception as e:
        return False, str(e)


def generate_pdf_report(columns, table_data, file_path, header_info, currency_columns=None, totals_row=None):
    """
    Generate a premium PDF report using reportlab with efficient styling.
    """
    try:
        from reportlab.lib.pagesizes import letter, landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm

        # Choose page size based on column count
        if len(columns) > 6:
            page_size = landscape(letter)
        else:
            page_size = landscape(A4)

        doc = SimpleDocTemplate(
            file_path, pagesize=page_size,
            rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20
        )

        elements = []
        styles = getSampleStyleSheet()

        # ── Color palette ──
        PRIMARY = colors.HexColor("#2F75B5")
        SECONDARY = colors.HexColor("#D9E1F2")
        DARK_TEXT = colors.HexColor("#1F2937")
        LIGHT_TEXT = colors.HexColor("#555555")
        ALT_ROW = colors.HexColor("#F2F2F2")
        BORDER_COLOR = colors.HexColor("#BFBFBF")

        # ── Styles ──
        company_style = ParagraphStyle(
            "Company", parent=styles["Heading1"],
            fontSize=14, textColor=PRIMARY, alignment=1, spaceAfter=2
        )
        sub_style = ParagraphStyle(
            "SubHeader", parent=styles["Normal"],
            fontSize=9, textColor=LIGHT_TEXT, alignment=1, spaceAfter=1
        )
        title_style = ParagraphStyle(
            "ReportTitle", parent=styles["Heading2"],
            fontSize=13, textColor=PRIMARY, alignment=1,
            spaceBefore=6, spaceAfter=2
        )
        date_style = ParagraphStyle(
            "DateRange", parent=styles["Normal"],
            fontSize=9, textColor=LIGHT_TEXT, alignment=1, spaceAfter=2
        )
        filter_style = ParagraphStyle(
            "FilterInfo", parent=styles["Normal"],
            fontSize=8, textColor=colors.HexColor("#888888"), alignment=1, spaceAfter=4
        )

        # ── Build header paragraphs ──
        elements.append(Paragraph(f"<b>{header_info.get('company_name', 'PharmIQ')}</b>", company_style))

        if header_info.get("address"):
            elements.append(Paragraph(header_info["address"], sub_style))

        contact_parts = []
        if header_info.get("phone"):
            contact_parts.append(f"Phone: {header_info['phone']}")
        if header_info.get("email"):
            contact_parts.append(f"Email: {header_info['email']}")
        if header_info.get("gst_no"):
            contact_parts.append(f"GST: {header_info['gst_no']}")
        if contact_parts:
            elements.append(Paragraph(" | ".join(contact_parts), sub_style))

        if header_info.get("licence_no"):
            elements.append(Paragraph(f"Licence No.: {header_info['licence_no']}", sub_style))

        elements.append(Spacer(1, 4 * mm))
        elements.append(Paragraph(f"<b>{header_info.get('report_title', 'Report')}</b>", title_style))
        elements.append(Paragraph(header_info.get("date_range", ""), date_style))

        # Filter info
        filter_parts = []
        if header_info.get("customer_filter"):
            filter_parts.append(f"Customer: {header_info['customer_filter']}")
        if header_info.get("medicine_filter"):
            filter_parts.append(f"Medicine: {header_info['medicine_filter']}")
        if filter_parts:
            elements.append(Paragraph(" | ".join(filter_parts), filter_style))

        elements.append(Spacer(1, 4 * mm))

        if not table_data:
            elements.append(Paragraph("No data found for the selected filters.", styles["Normal"]))
            doc.build(elements)
            return True, "Exported empty report"

        # ── Build table data (single pass string conversion) ──
        pdf_rows = [columns]  # Header row
        for row_dict in table_data:
            values = row_dict.get("values", [])
            pdf_rows.append([str(v) if v is not None else "" for v in values])

        # Add totals row
        if totals_row:
            pdf_rows.append([str(v) if v is not None else "" for v in totals_row])

        # ── Column widths ──
        page_width = page_size[0] - 40
        col_width = page_width / len(columns)

        t = Table(pdf_rows, colWidths=[col_width] * len(columns), repeatRows=1)

        # ── Table style (batch commands, not per-row) ──
        style_commands = [
            # Header row
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),

            # Data rows
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("TEXTCOLOR", (0, 1), (-1, -1), DARK_TEXT),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 1), (-1, -1), "LEFT"),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
            ("TOPPADDING", (0, 1), (-1, -1), 5),

            # Grid
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_COLOR),
            ("LINEBELOW", (0, 0), (-1, 0), 1.5, PRIMARY),
        ]

        # Alternating row colors (batch — just even rows)
        total_data_rows = len(table_data)
        for i in range(0, total_data_rows, 2):
            row_idx = i + 1  # +1 because header is row 0
            style_commands.append(("BACKGROUND", (0, row_idx), (-1, row_idx), ALT_ROW))

        # Totals row styling
        if totals_row:
            totals_idx = len(pdf_rows) - 1
            style_commands.extend([
                ("BACKGROUND", (0, totals_idx), (-1, totals_idx), SECONDARY),
                ("FONTNAME", (0, totals_idx), (-1, totals_idx), "Helvetica-Bold"),
                ("TEXTCOLOR", (0, totals_idx), (-1, totals_idx), PRIMARY),
                ("LINEABOVE", (0, totals_idx), (-1, totals_idx), 1.5, PRIMARY),
            ])

        # Right-align currency columns
        if currency_columns:
            for i, col in enumerate(columns):
                if col in currency_columns:
                    style_commands.append(("ALIGN", (i, 1), (i, -1), "RIGHT"))

        t.setStyle(TableStyle(style_commands))
        elements.append(t)

        # ── Footer line ──
        elements.append(Spacer(1, 6 * mm))
        footer_style = ParagraphStyle(
            "Footer", parent=styles["Normal"],
            fontSize=7, textColor=colors.HexColor("#AAAAAA"), alignment=1
        )
        elements.append(Paragraph(
            f"Generated by PharmIQ on {datetime.now().strftime('%Y-%m-%d %H:%M')} — This is a system generated report.",
            footer_style
        ))

        doc.build(elements)
        return True, "PDF report generated successfully."
    except Exception as e:
        return False, str(e)


# Required import for footer timestamp
from datetime import datetime
